import pandas as pds
import openpyxl
from openpyxl import load_workbook
import boto3
import os

s3 = boto3.resource('s3')

# Default bucket Name and FilePath
BucketName = "excelfile-pythonapp"
filePath = os.environ['FILE_PATH']
s3BucketFileName = "emp-data.xlsx"


file =(filePath)
data = pds.read_excel(file)

def uploadS3File():
    s3.Object(BucketName, s3BucketFileName).upload_file(filePath)

def readData():
    print("Read Mode \n")
    #reload data
    data = pds.read_excel(file)
    print(data)
    print('-----------------------------')

def writeData():
    print("Write mode \n")
    empId = int(input('Enter Emp id: '))
    empName = input('Enter emp name: ')
    jdt = int(input('Enter joining year: '))
    df = pds.DataFrame([[empId, empName, jdt]],
                       columns=['Employee Id', 'Employee Name', 'Year Of Joining'])
    ndata = data.append(df)
    with pds.ExcelWriter(filePath, mode="a", engine="openpyxl",if_sheet_exists='replace') as writer:
        ndata.to_excel(writer, sheet_name="Sheet1", index=False)
    print('-----------------------------')

def updateData():
    print("Update Mode Selected \n")
    empId = int(input('Enter employee Id to update its data: '))
    targetValue = str(input('Enter target value for update: '))
    workbook = load_workbook(filename=filePath)
    sheet = workbook.active
    for row in sheet.iter_rows():
        #get first object from row which is cell and then its row index
        rowIdx = row[0].row
        cell = sheet.cell(row=rowIdx,column=1)
        if cell.value == empId:
            targetCell = sheet.cell(row=rowIdx, column=2)
            targetCell.value=targetValue
    workbook.save(filePath)
    print('-----------------------------')

def deleteData():
    print("Delete Mode \n")
    empId = int(input('Enter employee Id whos data you want to delete: '))
    workbook = load_workbook(filename=filePath)
    sheet = workbook.active
    for row in sheet.iter_rows():
        rowIdx = row[0].row
        cell = sheet.cell(row=rowIdx,column=1)
        if cell.value == empId:
            sheet.delete_rows(rowIdx, 1)
    workbook.save(filePath)
    print('-----------------------------')
while True:
    print("Read Data from excel: 1")
    print("Write Data in excel: 2")
    print("Update Name by empId: 3")
    print("Delete data by empId: 4")
    print("Exit: 9")
    choice = int(input("Enter your choice "))
    if choice == 1:
        readData()

    if choice == 2:
        writeData()

    if choice == 3:
        updateData()

    if choice == 4:
        deleteData()

    if choice == 9:
        uploadS3File()
        break

print("Good bye")

